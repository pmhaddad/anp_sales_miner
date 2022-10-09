from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from anp_sales_miner.jobs.excel_parser import ExcelParser

FIXTURES_ROOT = '/var/anp_sales_miner/tests/fixtures'


@patch('anp_sales_miner.jobs.excel_parser.load_workbook')
def test_run(mock_load_workbook, monkeypatch, tmpdir):
    monkeypatch.setenv('DATALAKE_PATH', str(tmpdir))

    input_df = load_workbook(f'{FIXTURES_ROOT}/jobs/raw_oils_sales.xlsx')
    mock_load_workbook.return_value = input_df

    ExcelParser('Tabela dinâmica1', 'oils').run()

    output_fuel_df = pd.read_csv(f'{tmpdir}/parsed/oils.csv')
    output_totals_df = pd.read_csv(f'{tmpdir}/parsed/oils_totals.csv')

    expected_fuel_df = pd.read_csv(f'{FIXTURES_ROOT}/jobs/parsed_oils.csv')
    expected_totals_df = pd.read_csv(f'{FIXTURES_ROOT}/jobs/parsed_oils_totals.csv')

    assert (output_fuel_df.round(3) == expected_fuel_df.round(3)).all().all()

    assert (output_totals_df.round(3) == expected_totals_df.round(3)).all().all()
